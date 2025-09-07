import streamlit as st
from utils import *
import datetime
import io
import pandas as pd

# === Judul App ===
st.markdown(
    """
    <h1 style='
        text-align: center;
        font-size: 45px;
        font-weight: bold;
        background: -webkit-linear-gradient(45deg, #1ABC9C, #3498DB);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    '>
        Sistem Layanan Modifikasi Cuaca
    </h1>
    """,
    unsafe_allow_html=True
)
# === Form Input Data Baru ===
st.subheader("Layanan Jasa Survey Lokasi dan Comissioning")

# =================================================== #
# ==================== Baca Data ==================== #
# =================================================== #
df = pd.read_excel("database/data_gabungan_sdm2.xlsx", header=1) 

# ==================================================== #
# ==================== Input Data ==================== #
# ==================================================== #
today = datetime.date.today()

# hanya pilih tanggal mulai, bukan range
start_date = st.date_input(
    "Tanggal Mulai Pelaksanaan",
    today,
    min_value=today,
    format="DD.MM.YYYY",
)

# otomatis tentukan end_date supaya total 7 hari
end_date = start_date + datetime.timedelta(days=6)

# jumlah hari fix 7
jh_keseluruhan = 7

st.info(f"Pelaksanaan: {start_date} s/d {end_date} ({jh_keseluruhan} hari)")

# === Input Provinsi ===
provinsi = st.selectbox(
    "Tempat Pelaksanaan (Provinsi)",
    [
        "Aceh","Sumatera Utara","Sumatera Barat","Riau","Jambi","Sumatera Selatan",
        "Bengkulu","Lampung","Bangka Belitung","Kepulauan Riau","DKI Jakarta",
        "Jawa Barat","Jawa Tengah","DI Yogyakarta","Jawa Timur","Banten","Bali",
        "Nusa Tenggara Barat","Nusa Tenggara Timur","Kalimantan Barat","Kalimantan Tengah",
        "Kalimantan Selatan","Kalimantan Timur","Kalimantan Utara","Sulawesi Utara",
        "Sulawesi Tengah","Sulawesi Selatan","Sulawesi Tenggara","Gorontalo",
        "Sulawesi Barat","Maluku","Maluku Utara","Papua","Papua Barat","Papua Tengah",
        "Papua Pegunungan","Papua Selatan","Papua Barat Daya"
    ]
)

# === Input Jumlah Sistem Teleburning ===
jumlah_unit_sistem_teleburning = st.number_input(
    "Jumlah Sistem Teleburning",
    min_value=1, step=1, value=1
)

# ===================================================== #
# ==================== Variabel Fix =================== #
# ===================================================== #
# Jumlah Personel Lengkap
jumlah_personel_survey_lokasi = 7 # FIX, edit jika berubah
jumlah_personel_komisioning_instalasi = 7
jumlah_personel_penyusun_laporan_akhir = 7

# Ambil uang harian berdasarkan provinsi
row = df[df['PROVINSI'] == provinsi]
if row.empty:
    st.error(f"Provinsi '{provinsi}' tidak ditemukan di data Excel!")
    st.stop()

pp = 1

jumlah_lokasi = 1
jumlah_kali_lokasi = 1

jumlah_kegiatan = 1

jumlah_paket = 1
jumlah_kali_paket = 1

# ===================================================== #
# ==================== Uang Harian ==================== #
# ===================================================== #
uang_harian_luar_kota = row['luar_kota'].values[0]
uang_harian_dalam_kota = row['dalam_kota'].values[0]
uang_harian_khusus = df[df['PROVINSI'] == 'Jawa Barat']['luar_kota'].values[0]

jh_uh_survey_lokasi = jh_keseluruhan
jh_uh_komisioning_instalasi = 5
jh_uh_penyusun_laporan_akhir = 5

# Luar Kota
uh_survey_lokasi = hitung_uang_harian(uang_harian_luar_kota, jh_uh_survey_lokasi, jumlah_personel_survey_lokasi)
uh_komisioning_instalasi = hitung_uang_harian(uang_harian_luar_kota, jh_uh_komisioning_instalasi, jumlah_personel_komisioning_instalasi)
# Khusus
uh_penyusunan_laporan = hitung_uang_harian(uang_harian_khusus, jh_uh_penyusun_laporan_akhir, jumlah_personel_penyusun_laporan_akhir)

# ========================================================== #
# ==================== Biaya Penginapan ==================== #
# ========================================================== #
harga_penginapan_ktg3 = row['HOTELIII'].values[0]
harga_penginapan_ktg4 = row['HOTELIV'].values[0]
harga_penginapan_khusus = df[df['PROVINSI'] == 'Jawa Barat']['HOTELIV'].values[0]

jh_penginapan_survey_lokasi = jh_uh_survey_lokasi-1
jh_penginapan_komisioning_instalasi = jh_uh_komisioning_instalasi-1
jh_penginapan_penyusun_laporan_akhir = jh_uh_penyusun_laporan_akhir-1

# Biaya Penginapan Kategori III
biaya_penginapan_survey_lokasi = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_survey_lokasi, jumlah_personel_survey_lokasi)
biaya_penginapan_komisioning_instalasi = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_komisioning_instalasi, jumlah_personel_komisioning_instalasi)

# Biaya Penginapan Khusus
biaya_penginapan_penyusun_laporan_akhir = biaya_penginapan(harga_penginapan_khusus, jh_penginapan_penyusun_laporan_akhir, jumlah_personel_penyusun_laporan_akhir)

# ===================================================== #
# ==================== Biaya Tiket ==================== #
# ===================================================== #
biaya_tiket = row['PESAWAT_EKONOMI'].values[0] # Biaya tiket sama semua

biaya_tiket_survey_lokasi = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_survey_lokasi)
biaya_tiket_komisioning_instalasi = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_komisioning_instalasi)

# ===================================================== #
# ==================== Harga Taksi ==================== #
# ===================================================== #
harga_taksi_flat = 2 * df[df['PROVINSI'] == 'DKI Jakarta']['BANDARA'].values[0]

biaya_taksi_survey_lokasi = biaya_taksi(jumlah_personel_survey_lokasi, harga_taksi_flat)
biaya_taksi_komisioning_instalasi = biaya_taksi(jumlah_personel_komisioning_instalasi, harga_taksi_flat)
biaya_taksi_penyusun_laporan_akhir = biaya_taksi(jumlah_personel_penyusun_laporan_akhir, harga_taksi_flat)

# ===================================================================================== #
# ==================== Jasa Konsultasi Meteorologi dan Klimatologi ==================== #
# ===================================================================================== #
harga_jasa_konsultasi_meteorologi_klimatologi = 13250000

biaya_jasa_konsultasi_meteorologi_klimatologi = jasa_konsultasi_meteorologi_klimatologi(jumlah_lokasi, jumlah_kali_lokasi, harga_jasa_konsultasi_meteorologi_klimatologi)

# =============================================================================== #
# ==================== Jasa Konsultasi Kegiatan Perekayasaan ==================== #
# =============================================================================== #
harga_jasa_konsultasi_kegiatan_perekayasaan = 1540000

biaya_jasa_konsultasi_kegiatan_perekayasaan = jasa_konsultasi_kegiatan_perekayasaan(jumlah_kegiatan, harga_jasa_konsultasi_kegiatan_perekayasaan)

# ======================================================================= #
# ==================== Perolehan data cuaca historis ==================== #
# ======================================================================= #
# harga_perolehan_data_cuaca_historis = 20000000

# biaya_perolehan_data_cuaca_historis = perolehan_data_cuaca_historis(jumlah_paket, jumlah_kali_paket, harga_perolehan_data_cuaca_historis)

# ============================================================ #
# ==================== Sistem Teleburning ==================== #
# ============================================================ #
harga_sistem_teleburning = 25000000

biaya_sistem_teleburning = sistem_teleburning(jumlah_unit_sistem_teleburning, jumlah_kali_paket, harga_sistem_teleburning)

# ============================================= #
# ==================== Gas ==================== #
# ============================================= #
harga_gas = 5000000

biaya_gas = gas(jumlah_paket, jumlah_kali_paket, harga_gas)

# ===================================================== #
# ==================== Balon Pibal ==================== #
# ===================================================== #
harga_balon_pibal = 5000000

biaya_balon_pibal = balon_pibal(jumlah_paket, jumlah_kali_paket, harga_balon_pibal)

# ======================================================================================= #
# ==================== Bahan Semai Flare untuk Komisioning Instalasi Menara GBG ==================== #
# ============================================================================== #
harga_bahan_semai_flare = 2500000

jumlah_pcs, biaya_bahan_semai_flare = bahan_semai_flare(jumlah_unit_sistem_teleburning, 1, harga_bahan_semai_flare)

# ============================================================================================================ #
# ==================== Perijinan Bahan Semai Flare untuk Komisioning Instalasi Menara GBG ==================== #
# ============================================================================================================ #
harga_perijinan_bahan_semai_flare = 200000000

biaya_perijinan_bahan_semai_flare = perijinan_bahan_semai_flare(jumlah_paket, jumlah_kali_paket, harga_perijinan_bahan_semai_flare)

# ============================================================================ #
# ==================== biaya_sewa_kendaraan_survey_lokasi ==================== #
# ============================================================================ #
sewa_kendaraan = row['MOBIL'].values[0]
jumlah_unit_kendaraan_keseluruhan = math.ceil(jumlah_personel_survey_lokasi / 4)

biaya_sewa_kendaraan_survey_lokasi_GBG = biaya_sewa_kendaraan(jh_keseluruhan+3, 2, sewa_kendaraan) # pindah ke jasa_survey.py

# ============================================================================================== #
# ==================== Laporan Hasil Survey Lokasi dan Instalasi Menara GBG ==================== #
# ============================================================================================== #
harga_pencetakan_dan_penggandaan_laporan = 10000000
biaya_pencetakan_dan_penggandaan_laporan = pencetakan_dan_penggandaan_laporan(jumlah_paket, jumlah_kali_paket, harga_pencetakan_dan_penggandaan_laporan)

# === Debug / Tampilkan hasil ===
# st.write("📌 Ringkasan Input:")
# st.write("Jumlah Hari:", jh_keseluruhan)
# st.write("Provinsi:", provinsi)
# st.write("Jumlah Unit Sistem Teleburning:", jumlah_unit_sistem_teleburning)

# Buat tabel hasil
data = [
    ["Provinsi " + provinsi, "", "", "", "", "", ""],
    ["A. Personil Pelaksana", "", "", "", "", "", ""],

    ["1. Survey Lokasi Wahana Penyemai Awan dari Darat", "", "", "", "", "", ""],
    ["     Uang Harian", jumlah_personel_survey_lokasi, "orang", jh_uh_survey_lokasi, "hari", uang_harian_luar_kota, uh_survey_lokasi],
    ["     Biaya Penginapan", jumlah_personel_survey_lokasi, "orang", jh_penginapan_survey_lokasi, "hari", harga_penginapan_ktg4, biaya_penginapan_survey_lokasi],
    ["     Biaya Tiket ", jumlah_personel_survey_lokasi, "orang", pp, "pp", biaya_tiket, biaya_tiket_survey_lokasi],
    ["     Taksi", jumlah_personel_survey_lokasi, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_survey_lokasi],
    ["     Jasa Konsultasi Meteorologi dan Klimatologi", jumlah_lokasi, "lokasi", jumlah_kali_lokasi, "kali", harga_jasa_konsultasi_meteorologi_klimatologi, biaya_jasa_konsultasi_meteorologi_klimatologi],

    ["2. Commissioning Instalasi Wahana Penyemai Awan Dari Darat", "", "", "", "", "", ""],
    ["     Uang Harian", jumlah_personel_komisioning_instalasi, "orang", jh_uh_komisioning_instalasi, "hari", uang_harian_luar_kota, uh_komisioning_instalasi],
    ["     Biaya Penginapan", jumlah_personel_komisioning_instalasi, "orang", jh_penginapan_komisioning_instalasi, "hari", harga_penginapan_ktg4, biaya_penginapan_komisioning_instalasi],
    ["     Biaya Tiket ", jumlah_personel_komisioning_instalasi, "orang", pp, "pp", biaya_tiket, biaya_tiket_komisioning_instalasi],
    ["     Taksi", jumlah_personel_komisioning_instalasi, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_komisioning_instalasi],
    ["     Jasa Konsultasi Kegiatan Perekayasaan", jumlah_kegiatan, "kegiatan", "", "", harga_jasa_konsultasi_kegiatan_perekayasaan, biaya_jasa_konsultasi_kegiatan_perekayasaan],

    ["3. Penyusunan Laporan", "", "", "", "", "", ""],
    ["     Uang Harian", jumlah_personel_penyusun_laporan_akhir, "orang", jh_uh_penyusun_laporan_akhir, "hari", uang_harian_luar_kota, uh_penyusunan_laporan],
    ["     Biaya Penginapan", jumlah_personel_penyusun_laporan_akhir, "orang", jh_penginapan_penyusun_laporan_akhir, "hari", harga_penginapan_khusus, biaya_penginapan_penyusun_laporan_akhir],
    ["     Taksi", jumlah_personel_penyusun_laporan_akhir, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_penyusun_laporan_akhir],

    ["B. Sarana dan Prasarana", "", "", "", "", "", ""],
    ["      Sistem Teleburning", jumlah_unit_sistem_teleburning, "unit", jumlah_kali_paket, "kali", harga_sistem_teleburning, biaya_sistem_teleburning],
    ["      Balon Pibal", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_balon_pibal, biaya_balon_pibal],
    ["      Bahan Semai Flare untuk komisioning instalasi Menara GBG", jumlah_pcs, "pcs", jumlah_kali_paket, "hari", harga_bahan_semai_flare, biaya_bahan_semai_flare],
    ["      Perijinan Bahan Semai Flare untuk komisioning instalasi Menara GBG", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_perijinan_bahan_semai_flare, biaya_perijinan_bahan_semai_flare],
    ["      Sewa Kendaraan Survey Lokasi dan Komisioning Instalasi Menara GBG", jumlah_unit_kendaraan_keseluruhan, "unit", jh_keseluruhan+3, "hari", sewa_kendaraan, biaya_sewa_kendaraan_survey_lokasi_GBG],

    ["C. Hasil Survey Lokasi dan Commissioning Instalasi", "", "", "", "", "", ""],
    ["Laporan pelaksanaan hasil survei", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_pencetakan_dan_penggandaan_laporan, biaya_pencetakan_dan_penggandaan_laporan],
]

jumlah_total_biaya_survey_lokasi = sum(row[-1] for row in data if row[-1] != "")


df_out = pd.DataFrame(data, columns=[
    "Uraian", "Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume", "Indeks (Rupiah)", "Biaya (Rupiah)"
])


jumlah_harian = jumlah_total_biaya_survey_lokasi / jh_keseluruhan

# buat baris ringkasan
summary_rows = pd.DataFrame([
    ["Jumlah", "", "", "", "", f"Total biaya {jh_keseluruhan} hari", jumlah_total_biaya_survey_lokasi],
    ["", "", "", "", "", "Tarif Harian", jumlah_harian]
], columns=df_out.columns)

# gabungkan tabel utama dengan ringkasan
df_final = pd.concat([df_out, summary_rows], ignore_index=True)

# format angka pakai koma ribuan
# pastikan dulu dua kolom angka jadi numeric
df_final["Indeks (Rupiah)"] = pd.to_numeric(df_final["Indeks (Rupiah)"], errors="coerce")
df_final["Biaya (Rupiah)"]  = pd.to_numeric(df_final["Biaya (Rupiah)"], errors="coerce")

# lalu format dengan ribuan pakai koma
df_final["Indeks (Rupiah)"] = df_final["Indeks (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)
df_final["Biaya (Rupiah)"] = df_final["Biaya (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)

# st.write("📌 Rincian Biaya:", jumlah_total_biaya_survey_lokasi)

# tampilkan dataframe di streamlit
# tampilkan tabel dengan styling
st.markdown(
    """
    <style>
    .dataframe td, .dataframe th {
        white-space: nowrap;
        text-align: left;
        padding: 8px 12px;
    }
    .dataframe th {
        background-color: #f5f5f5;
    }
    .dataframe td:nth-child(1) { min-width: 300px; }  /* kolom Uraian */
    .dataframe td:nth-child(2) { min-width: 50px; text-align: right; } /* kolom Jumlah */
    .dataframe td:nth-child(7) { min-width: 120px; text-align: right; } /* kolom Biaya */
    </style>
    """,
    unsafe_allow_html=True
)

st.dataframe(df_final, use_container_width=True)

from openpyxl.styles import Alignment, Font, numbers, Border, Side

# simpan ke buffer, bukan file fisik
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Laporan")

    ws = writer.sheets["Laporan"]

    # auto-width kolom
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # header bold & rata tengah + border
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border
        # semua header rata tengah
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- isi tabel ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value  # ambil nama header
            if col_name in ["Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

# reset pointer buffer
buffer.seek(0)

# tombol download
st.download_button(
    label="📥 Download Excel",
    data=buffer,
    file_name=f"Layanan Jasa Survey Lokasi dan Comissioning {provinsi}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
st.markdown("---", unsafe_allow_html=True)
st.markdown(
    """
    <style>
        .footer-container {
            background-color: #f0f2f6;
            padding: 20px 40px;
            font-size: 14px;
            color: #333;
            display: flex;
            flex-wrap: wrap;
            justify-content: space-between;
        }
        .footer-column {
            flex: 1;
            min-width: 200px;
            margin-bottom: 20px;
        }
        .footer-title {
            font-weight: bold;
            margin-bottom: 10px;
            color: #0059b3;
        }
        .footer-bottom {
            text-align: center;
            font-size: 13px;
            padding: 10px;
            color: gray;
            border-top: 1px solid #ccc;
            background-color: #f0f2f6;
        }
    </style>

    <div class="footer-container">
        <div class="footer-column">
            <div class="footer-title">Kontak Kami</div>
            Jl. Angkasa I No.2 Kemayoran,<br>
            Jakarta Pusat 10610,<br>
            PO Box 3540 Jkt.<br><br>
            Contact Center: (021) 4246321<br>
            Faks: (021) 4246703<br>
            Email: tu5@bmkg.go.id
        </div>
        <div class="footer-column">
            <div class="footer-title">Media Sosial</div>
            Instagram : dmc_bmkg<br>
        </div>
    </div>

    <div class="footer-bottom">
        © 2025 - Sub Bagian Tata Usaha Modifikasi Cuaca, Deputi Bidang Modifikasi Cuaca, BMKG.
    </div>
    """,
    unsafe_allow_html=True
)
