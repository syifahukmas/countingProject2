import streamlit as st
from utils import *
import datetime
import io
import pandas as pd

# === Judul App ===
st.markdown(
    """
    <h1 style='
        text-align: center;
        font-size: 45px;
        font-weight: bold;
        background: -webkit-linear-gradient(45deg, #1ABC9C, #3498DB);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    '>
        Sistem Layanan Modifikasi Cuaca
    </h1>
    """,
    unsafe_allow_html=True
)
# === Form Input Data Baru ===
st.subheader("Layanan Jasa Operasi Modifikasi Cuaca")

# =================================================== #
# ==================== Baca Data ==================== #
# =================================================== #
df = pd.read_excel("database/data_gabungan_sdm2.xlsx", header=1) 

# ==================================================== #
# ==================== Input Data ==================== #
# ==================================================== #
# today = datetime.date.today()
# tanggal_range = st.date_input(
#     "Tanggal Pelaksanaan",
#     (today, today + datetime.timedelta(days=6)),  # default 7 hari
#     min_value=today,
#     format="DD.MM.YYYY",
# )

# # Hitung jumlah hari
# jh_keseluruhan = None
# if isinstance(tanggal_range, tuple) and len(tanggal_range) == 2:
#     start_date, end_date = tanggal_range
#     jh_keseluruhan = (end_date - start_date).days + 1
#     st.info(f"Pelaksanaan: {start_date} s/d {end_date} ({jh_keseluruhan} hari)")

today = datetime.date.today()

# hanya pilih tanggal mulai, bukan range
start_date = st.date_input(
    "Tanggal Mulai Pelaksanaan ",
    today,
    min_value=today,
    format="DD.MM.YYYY",
)
jumlah_input_hari=st.number_input("Masukkan Jumlah Hari :  ", step=1)
# otomatis tentukan end_date supaya total 7 hari
end_date = start_date + datetime.timedelta(days=jumlah_input_hari)

jh_keseluruhan=jumlah_input_hari

st.info(f"Pelaksanaan: {start_date} s/d {end_date} ({jh_keseluruhan} hari)")
# === Input Provinsi ===
provinsi = st.selectbox(
    "Tempat Pelaksanaan (Provinsi)",
    [
        "Aceh","Sumatera Utara","Sumatera Barat","Riau","Jambi","Sumatera Selatan",
        "Bengkulu","Lampung","Bangka Belitung","Kepulauan Riau","DKI Jakarta",
        "Jawa Barat","Jawa Tengah","DI Yogyakarta","Jawa Timur","Banten","Bali",
        "Nusa Tenggara Barat","Nusa Tenggara Timur","Kalimantan Barat","Kalimantan Tengah",
        "Kalimantan Selatan","Kalimantan Timur","Kalimantan Utara","Sulawesi Utara",
        "Sulawesi Tengah","Sulawesi Selatan","Sulawesi Tenggara","Gorontalo",
        "Sulawesi Barat","Maluku","Maluku Utara","Papua","Papua Barat","Papua Tengah",
        "Papua Pegunungan","Papua Selatan","Papua Barat Daya"
    ]
)

# === Pilih Jenis Pesawat ===
jenis_pesawat = st.selectbox(
    "Pilih Jenis Pesawat",
    ["", "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)", "Sewa Pesawat Swasta"]  # "" supaya default kosong
)

# === Munculkan input sesuai pilihan pesawat ===
jenis_operasi = None
jumlah_personel_keseluruhan = None
jumlah_kru_pesawat_TNI = None

if jenis_pesawat == "Sewa Pesawat Swasta":
    # Pilih Jenis Operasi
    jenis_operasi = st.selectbox(
        "Jenis Operasi",
        ["Kondisi Normal (Operasi 12 Jam)", "Kondisi Tertentu/Khusus (Operasi 24 Jam)"]
    )
    # Input Jumlah Personel OMC
    jumlah_personel_keseluruhan = st.number_input(
        "Jumlah Personel OMC (minimal 40)",
        min_value=40, step=1, value=40
    )

elif jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)":
    # Input Jumlah Personel OMC
    jumlah_personel_keseluruhan = st.number_input(
        "Jumlah Personel OMC (minimal 40)",
        min_value=40, step=1, value=40
    )
    # Input Jumlah Personel TNI
    jumlah_kru_pesawat_TNI = st.number_input(
        "Jumlah Personel TNI (11-16)",
        min_value=11, max_value=16, step=1, value=11
    )
    

# ===================================================== #
# ==================== Variabel Fix =================== #
# ===================================================== #
# Jumlah Personel Lengkap
jumlah_personel_tim_pengawas = 3 # FIX, edit jika berubah
jumlah_personel_supervisi_pimpinan = 3
jumlah_personel_observer_cuaca = 6
jumlah_personel_pendukung_posko = 6
jumlah_personel_tim_pelaksana = safe_subtract(
    jumlah_personel_keseluruhan,
    jumlah_personel_tim_pengawas,
    jumlah_personel_supervisi_pimpinan,
    jumlah_personel_observer_cuaca,
    jumlah_personel_pendukung_posko
)

# Formasi Personel Pelengkap
jumlah_personel_koordinasi_awal = 6 # FIX, edit jika berubah
jumlah_personel_instalasi_posko = 2
# jumlah_personel_supervisi_pimpinan = 3
jumlah_personel_uninstall_posko = 2
jumlah_personel_pendukung_posko = 6
jumlah_personel_penyusun_laporan_akhir = 12
jumlah_personel_dukungan_teknis_operasional = 65

# Input Jumlah Hari Kegiatan
# jh = jumlah_hari
# uh = uang harian
jh_uh_koordinasi_awal = 3
jh_uh_instalasi_posko = 3
jh_uh_tim_pelaksana = jh_keseluruhan + 2
jh_uh_kru_pesawat = jh_keseluruhan + 2
jh_uh_tim_pengawas = 5
jh_uh_supervisi_pimpinan = 5
jh_uh_uninstal_posko = 3
jh_uh_penyusunan_laporan_akhir = 5

# Jumlah Paket Secara Umum
jumlah_paket = 1

# Jumlah Kali Paket
jumlah_kali_paket = 1

# Ambil uang harian berdasarkan provinsi
row = df[df['PROVINSI'] == provinsi]
if row.empty:
    st.error(f"Provinsi '{provinsi}' tidak ditemukan di data Excel!")
    st.stop()
def harga_sewa_pesawat(input_operasi, jh_keseluruhan, harga_sewa_per_jam):
    if input_operasi == "Kondisi Tertentu/Khusus (Operasi 24 Jam)":
        jam_terbang = 6
    else:
        jam_terbang = 3
    jumlah_alutsista_pesawat = jam_terbang * jh_keseluruhan * harga_sewa_per_jam
    return jam_terbang, jumlah_alutsista_pesawat

# ===================================================== #
# ==================== Uang Harian ==================== #
# ===================================================== #
uang_harian_luar_kota = row['luar_kota'].values[0]
uang_harian_dalam_kota = row['dalam_kota'].values[0]
uang_harian_khusus = df[df['PROVINSI'] == 'Jawa Barat']['luar_kota'].values[0]

# Luar Kota
uh_koordinasi_awal = hitung_uang_harian(uang_harian_luar_kota, jh_uh_koordinasi_awal, jumlah_personel_koordinasi_awal)
uh_instalasi_posko = hitung_uang_harian(uang_harian_luar_kota, jh_uh_instalasi_posko, jumlah_personel_instalasi_posko)
uh_tim_pelaksana = hitung_uang_harian(uang_harian_luar_kota, jh_uh_tim_pelaksana, jumlah_personel_tim_pelaksana)

if jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)":
    uh_kru_pesawat = hitung_uang_harian(uang_harian_luar_kota, jh_uh_kru_pesawat, jumlah_kru_pesawat_TNI)
else:
    uh_kru_pesawat = 0

uh_tim_pengawas = hitung_uang_harian(uang_harian_luar_kota, jh_uh_tim_pengawas, jumlah_personel_tim_pengawas)
uh_supervisi_pimpinan = hitung_uang_harian(uang_harian_luar_kota, jh_uh_supervisi_pimpinan, jumlah_personel_supervisi_pimpinan)
uh_uninstal_posko = hitung_uang_harian(uang_harian_luar_kota, jh_uh_uninstal_posko, jumlah_personel_uninstall_posko)

# Dalam Kota
uh_observer_cuaca = hitung_uang_harian(uang_harian_dalam_kota, jh_keseluruhan, jumlah_personel_observer_cuaca)
uh_personal_pendukung_posko = hitung_uang_harian(uang_harian_dalam_kota, jh_keseluruhan, jumlah_personel_pendukung_posko)

# Khusus
uh_penyusunan_laporan = hitung_uang_harian(uang_harian_khusus, jh_uh_penyusunan_laporan_akhir, jumlah_personel_penyusun_laporan_akhir)
# ========================================================== #
# ==================== Biaya Penginapan ==================== #
# ========================================================== #

# jh = jumlah_hari
jh_penginapan_koordinasi_awal = jh_uh_koordinasi_awal - 1
jh_penginapan_instalasi_posko = jh_uh_instalasi_posko - 1
jh_penginapan_tim_pelaksana = jh_uh_tim_pelaksana - 1
jh_penginapan_kru_pesawat = jh_uh_kru_pesawat - 1
jh_penginapan_tim_pengawas = jh_uh_tim_pengawas - 1
jh_penginapan_supervisi_pimpinan = jh_uh_supervisi_pimpinan - 1
jh_penginapan_uninstal_posko = jh_uh_uninstal_posko - 1
jh_penginapan_penyusunan_laporan_akhir = jh_uh_penyusunan_laporan_akhir - 1

harga_penginapan_ktg3 = row['HOTELIII'].values[0]
harga_penginapan_ktg4 = row['HOTELIV'].values[0]
harga_penginapan_khusus = df[df['PROVINSI'] == 'Jawa Barat']['HOTELIV'].values[0]

# Biaya Penginapan Kategori III
# salah
biaya_penginapan_koordinasi_awal = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_koordinasi_awal, jumlah_personel_koordinasi_awal)
biaya_penginapan_instalasi_posko = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_instalasi_posko, jumlah_personel_instalasi_posko)
biaya_penginapan_tim_pelaksana = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_tim_pelaksana, jumlah_personel_tim_pelaksana)
if jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)"
    biaya_penginapan_kru_pesawat = biaya_penginapan(
        harga_penginapan_ktg4, jh_penginapan_kru_pesawat, jumlah_kru_pesawat_TNI
    )
else:
    biaya_penginapan_kru_pesawat = 0

# bener
biaya_penginapan_tim_pengawas = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_tim_pengawas, jumlah_personel_tim_pengawas)
biaya_penginapan_supervisi_pimpinan = biaya_penginapan(harga_penginapan_ktg3, jh_penginapan_supervisi_pimpinan, jumlah_personel_supervisi_pimpinan)

# Biaya Penginapan Kategori IV
# salah
biaya_penginapan_uninstal_posko = biaya_penginapan(harga_penginapan_ktg4, jh_penginapan_uninstal_posko, jumlah_personel_uninstall_posko)

# Biaya Penginapan Khusus
# bener
biaya_penginapan_penyusunan_laporan_akhir = biaya_penginapan(harga_penginapan_khusus, jh_penginapan_penyusunan_laporan_akhir, jumlah_personel_penyusun_laporan_akhir)

# ===================================================== #
# ==================== Biaya Tiket ==================== #
# ===================================================== #

pp = 1
p = 1
biaya_tiket = row['PESAWAT_EKONOMI'].values[0] # Biaya tiket sama semua

biaya_tiket_koordinasi_awal = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_koordinasi_awal)
biaya_tiket_instalasi_posko = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_instalasi_posko)
biaya_tiket_tim_pelaksana = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_tim_pelaksana)
biaya_tiket_tim_pengawas = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_tim_pengawas)
biaya_tiket_supervisi_pimpinan = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_supervisi_pimpinan)
biaya_tiket_uninstall_posko = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_uninstall_posko)

# ===================================================== #
# ==================== Harga Taksi ==================== #
# ===================================================== #

# jasa_omc.py
harga_taksi_flat = 2 * df[df['PROVINSI'] == 'DKI Jakarta']['BANDARA'].values[0]

biaya_taksi_koordinasi_awal = biaya_taksi(jumlah_personel_koordinasi_awal, harga_taksi_flat)
biaya_taksi_instalasi_posko = biaya_taksi(jumlah_personel_instalasi_posko, harga_taksi_flat)
biaya_taksi_tim_pelaksana = biaya_taksi(jumlah_personel_tim_pelaksana, harga_taksi_flat)
if jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)":
    biaya_taksi_kru_pesawat = biaya_taksi(jumlah_kru_pesawat_TNI, harga_taksi_flat)
else:
    biaya_taksi_kru_pesawat = 0
biaya_taksi_tim_pengawas = biaya_taksi(jumlah_personel_tim_pengawas, harga_taksi_flat)
biaya_taksi_supervisi_pimpinan = biaya_taksi(jumlah_personel_supervisi_pimpinan, harga_taksi_flat)
biaya_taksi_uninstal_posko = biaya_taksi(jumlah_personel_uninstall_posko, harga_taksi_flat)
biaya_taksi_penyusunan_laporan_akhir = biaya_taksi(jumlah_personel_penyusun_laporan_akhir, harga_taksi_flat)

# ========================================================== #
# ==================== Sarana Prasarana ==================== #
# ========================================================== #

# =================== Bahan Semai (NaCl) ==================== #
jumlah_kg = 1600
harga_per_kg = 14000
biaya_harga_bahan_semai_NaCl = harga_bahan_semai_NaCl(jumlah_kg, harga_per_kg, jh_keseluruhan)

# =================== Alutsista ==================== #
# =========== Penggantian Avtur Pesawat ============ #
jumlah_liter = 850
harga_avtur_per_liter = 16000
jam_mobdemob = row['jam_avtur_mobdemob'].values[0]
biaya_penggantian_avtur_pesawat = penggantian_avtur_pesawat(jumlah_liter, harga_avtur_per_liter, jam_mobdemob)

# # =========== Selama Operasi ============ #
mode = jenis_operasi
if jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)":
    jam_terbang_operasi, biaya_alutsista_pesawat_selama_operasi = alutsista_pesawat_selama_operasi(
        mode=="Kondisi Normal (Operasi 12 Jam)", jh_keseluruhan, jumlah_liter, harga_avtur_per_liter)
else:
    jam_terbang_operasi, biaya_alutsista_pesawat_selama_operasi = alutsista_pesawat_selama_operasi(
        mode, jh_keseluruhan, jumlah_liter, harga_avtur_per_liter)

jam_terbang_total = jam_terbang_operasi * jh_keseluruhan
# Modifikasi dan inspeksi pesawat before - after rain making 
harga_paket_modifikasi_inspeksi = 140000000
biaya_modifikasi_dan_inspeksi_pesawat = modifikasi_dan_inspeksi_pesawat(harga_paket_modifikasi_inspeksi, jumlah_paket, jumlah_kali_paket)

# ================ Biaya Sewa Per jam (Pesawat Swasta) ================ #
harga_sewa_per_jam = 70000000
jam_terbang_operasi, biaya_harga_sewa_pesawat = harga_sewa_pesawat(jenis_operasi, jh_keseluruhan, harga_sewa_per_jam)

harga_mobdemob = harga_sewa_per_jam * jam_mobdemob
# =================== Kebutuhan Operasional Lapangan ==================== #
# ========== Sewa Kendaraan ========== #
# Variabel sewa kendaraan
sewa_kendaraan = row['MOBIL'].values[0]
jumlah_unit_kendaraan_keseluruhan = math.ceil(safe_divide(jumlah_personel_keseluruhan, 4))
# Jumlah unit kendaraan
unit_kendaraan_sebelum_operasi = 2 
unit_kendaraan_setelah_operasi = 1
# Jumlah hari
jh_sewa_kendaraan_sebelum_setelah_operasi = 3
# Biaya Sewa
biaya_sewa_kendaraan_sebelum_operasi = biaya_sewa_kendaraan(jh_sewa_kendaraan_sebelum_setelah_operasi, unit_kendaraan_sebelum_operasi, sewa_kendaraan)
biaya_sewa_kendaraan_setelah_operasi = biaya_sewa_kendaraan(jh_sewa_kendaraan_sebelum_setelah_operasi, unit_kendaraan_setelah_operasi, sewa_kendaraan)
biaya_sewa_kendaraan_selama_operasi = biaya_sewa_kendaraan(jh_keseluruhan, jumlah_unit_kendaraan_keseluruhan, sewa_kendaraan)

# ========== Peralatan dan Pendukung lapangan (ATK dan  Rompi Safety) ========== #
harga_peralatan_dan_pendukung_lapangan = 4000000
biaya_peralatan_dan_pendukung_lapangan = peralatan_dan_pendukung_lapangan(jh_keseluruhan, jumlah_paket, harga_peralatan_dan_pendukung_lapangan)

# ========== Sewa Ruangan untuk Posko dan Gudang Bahan Semai ========== #
harga_paket_sewa_gedung = 3000000
biaya_sewa_gedung = sewa_gedung(jh_keseluruhan, jumlah_paket, harga_paket_sewa_gedung)

# ========== Ekspedisi ========== #
# Ekspedisi Peralatan

biaya_kirim = row['pengiriman_bahan_semai'].values[0]

total_berat_peralatan = 3000 # 
# biaya_per_kg_peralatan = 4000 # Tergantung daerahnya

total_berat_bahan_semai = 1600 * jh_keseluruhan
# biaya_per_kg_bahan_semai = 4000 # Tergantung daerahnya

biaya_ekspedisi_peralatan = biaya_ekspedisi(total_berat_peralatan, biaya_kirim, 2)
biaya_ekspedisi_bahan_semai = biaya_ekspedisi(total_berat_bahan_semai, biaya_kirim, p)

# ========== Dukungan teknis operasional di lapangan ========== #
uang_harian_dukungan_teknis = row['konsumsi_rapat'].values[0]
biaya_dukungan_teknis_operasional_di_lapangan = dukungan_teknis_operasional_di_lapangan(uang_harian_dukungan_teknis, jh_keseluruhan, jumlah_personel_dukungan_teknis_operasional)

# ========== Stanby  Air mobil  Pemadam Kebakaran dan untuk Cuci Pesawat ========== #
harga_air_mobil_pk_cuci_pesawat = 1000000
biaya_air_mobil_pk_cuci_pesawat = air_mobil_pk_cuci_pesawat(jumlah_paket, jh_keseluruhan, harga_air_mobil_pk_cuci_pesawat)

# ========== Groundhandlng ( Forklift, Ground Support Equipment) ========== #
harga_groundhandling = 4000000
biaya_groundhandling = groundhandling(jumlah_paket, jh_keseluruhan, harga_groundhandling)

# ========== C. Pelaporan ========== #
# ========== Pencetakan dan Penggandaan Laporan ========== #
harga_pencetakan_dan_penggandaan_laporan = 10000000
biaya_pencetakan_dan_penggandaan_laporan = pencetakan_dan_penggandaan_laporan(jumlah_paket, jumlah_kali_paket, harga_pencetakan_dan_penggandaan_laporan)

# Buat tabel hasil
data = [
    ["Provinsi " + provinsi, "", "", "", "", "", ""],
    ["A. Personil Pelaksana", "", "", "", "", "", ""],

    ["1. Sebelum Operasi", "", "", "", "", "", ""],
    ["1.1. Koordinasi Awal, Perizinan dan Administrasi Kontrak", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_koordinasi_awal, "orang", jh_uh_koordinasi_awal, "hari", uang_harian_luar_kota, uh_koordinasi_awal],
    ["Biaya Penginapan", jumlah_personel_koordinasi_awal, "orang", jh_penginapan_koordinasi_awal, "hari", harga_penginapan_ktg4, biaya_penginapan_koordinasi_awal],
    ["Biaya Tiket", jumlah_personel_koordinasi_awal, "orang", pp, "pp", biaya_tiket, biaya_tiket_koordinasi_awal],
    ["Taksi", jumlah_personel_koordinasi_awal, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_koordinasi_awal],
    
    ["1.2. Instalasi Posko", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_instalasi_posko, "orang", jh_uh_instalasi_posko, "hari", uang_harian_luar_kota, uh_instalasi_posko],
    ["Biaya Penginapan", jumlah_personel_instalasi_posko, "orang", jh_penginapan_instalasi_posko, "hari", harga_penginapan_ktg4, biaya_penginapan_instalasi_posko],
    ["Biaya Tiket", jumlah_personel_instalasi_posko, "orang", pp, "pp", biaya_tiket, biaya_tiket_instalasi_posko],
    ["Taksi", jumlah_personel_instalasi_posko, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_instalasi_posko],

    ["2. Selama Operasi", "", "", "", "", "", ""],
    ["2.1. Tim Pelaksana Posko", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_tim_pelaksana, "orang", jh_uh_tim_pelaksana, "hari", uang_harian_luar_kota, uh_tim_pelaksana],
    ["Biaya Penginapan", jumlah_personel_tim_pelaksana, "orang", jh_penginapan_tim_pelaksana, "hari", harga_penginapan_ktg4, biaya_penginapan_tim_pelaksana],
    ["Biaya Tiket", jumlah_personel_tim_pelaksana, "orang", pp, "pp", biaya_tiket, biaya_tiket_tim_pelaksana],
    ["Taksi", jumlah_personel_tim_pelaksana, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_tim_pelaksana],
]

# tambahkan Kru Pesawat TNI hanya jika pilihannya "Pesawat TNI"
if jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)":
    data += [
        ["2.2. Kru Pesawat TNI", "", "", "", "", "", ""],
        ["Uang Harian", jumlah_kru_pesawat_TNI, "orang", jh_uh_kru_pesawat, "hari", uang_harian_luar_kota, uh_kru_pesawat],
        ["Biaya Penginapan", jumlah_kru_pesawat_TNI, "orang", jh_penginapan_kru_pesawat, "hari", harga_penginapan_ktg4, biaya_penginapan_kru_pesawat],
        ["Taksi", jumlah_kru_pesawat_TNI, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_kru_pesawat],
    ]

# lanjutkan dengan Tim Pengawas, Supervisi, dll...
# data += [
#     ["2.3. Tim Pengawas (Inspektorat/APIP), Humas", "", "", "", "", "", ""],
#     ["Uang Harian", jumlah_personel_tim_pengawas, "orang", jh_uh_tim_pengawas, "hari", uang_harian_luar_kota, uh_tim_pengawas],
#     ["Biaya Penginapan", jumlah_personel_tim_pengawas, "orang", jh_penginapan_tim_pengawas, "hari", harga_penginapan_ktg4, biaya_penginapan_tim_pengawas],
#     ["Biaya Tiket", jumlah_personel_tim_pengawas, "orang", pp, "pp", biaya_tiket, biaya_tiket_tim_pengawas],
#     ["Taksi", jumlah_personel_tim_pengawas, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_tim_pengawas],
# ]

data += [
    ["2.3. Tim Pengawas (Inspektorat/APIP), Humas", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_tim_pengawas, "orang", jh_uh_tim_pengawas, "hari", uang_harian_luar_kota, uh_tim_pengawas],
    ["Biaya Penginapan", jumlah_personel_tim_pengawas, "orang", jh_penginapan_tim_pengawas, "hari", harga_penginapan_ktg4, biaya_penginapan_tim_pengawas],
    ["Biaya Tiket", jumlah_personel_tim_pengawas, "orang", pp, "pp", biaya_tiket, biaya_tiket_tim_pengawas],
    ["Taksi", jumlah_personel_tim_pengawas, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_tim_pengawas],

    ["2.4. Supervisi Pimpinan", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_supervisi_pimpinan, "orang", jh_uh_supervisi_pimpinan, "hari", uang_harian_luar_kota, uh_supervisi_pimpinan],
    ["Biaya Penginapan", jumlah_personel_supervisi_pimpinan, "orang", jh_penginapan_supervisi_pimpinan, "hari", harga_penginapan_ktg4, biaya_penginapan_supervisi_pimpinan],
    ["Biaya Tiket", jumlah_personel_supervisi_pimpinan, "orang", pp, "pp", biaya_tiket, biaya_tiket_supervisi_pimpinan],
    ["Taksi", jumlah_personel_supervisi_pimpinan, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_supervisi_pimpinan],

    ["2.5. Observer cuaca/UPT Daerah", "", "", "", "", "", ""],
    ["Uang Harian Dalam Kota", jumlah_personel_observer_cuaca, "orang", jh_keseluruhan, "hari", uang_harian_dalam_kota, uh_observer_cuaca],
    
    ["2.6. Personil Pendukung Posko (Lanud/Angkasa Pura)", "", "", "", "", "", ""],
    ["Uang Harian Dalam Kota", jumlah_personel_pendukung_posko, "orang", jh_keseluruhan, "hari", uang_harian_dalam_kota, uh_personal_pendukung_posko],

    ["3. Setelah Operasi", "", "", "", "", "", ""],
    ["3.1. Unistall Posko", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_uninstall_posko, "orang", jh_uh_uninstal_posko, "hari", uang_harian_luar_kota, uh_uninstal_posko],
    ["Biaya Penginapan", jumlah_personel_uninstall_posko, "orang", jh_penginapan_uninstal_posko, "hari", harga_penginapan_ktg4, biaya_penginapan_uninstal_posko],
    ["Biaya Tiket", jumlah_personel_uninstall_posko, "orang", pp, "pp", biaya_tiket, biaya_tiket_uninstall_posko],
    ["Taksi", jumlah_personel_uninstall_posko, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_uninstal_posko],

    ["3.2. Penyusunan Laporan Akhir Kegiatan", "", "", "", "", "", ""],
    ["Uang Harian", jumlah_personel_penyusun_laporan_akhir, "orang", jh_uh_penyusunan_laporan_akhir, "hari", uang_harian_khusus, uh_penyusunan_laporan],
    ["Biaya Penginapan", jumlah_personel_penyusun_laporan_akhir, "orang", jh_penginapan_penyusunan_laporan_akhir, "hari", harga_penginapan_ktg4, biaya_penginapan_penyusunan_laporan_akhir],
    ["Taksi", jumlah_personel_penyusun_laporan_akhir, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_penyusunan_laporan_akhir],

    ["B. Sarana dan Prasarana", "", "", "", "", "", ""],
    ["1. Bahan Semai", jumlah_kg, "Kg", jh_keseluruhan, "hari", harga_per_kg, biaya_harga_bahan_semai_NaCl],
]

if jenis_pesawat == "Pesawat TNI (Kebencanaan/Strategis Kenegaraan)":
    data += [
        ["2. Alutsista Pesawat TNI", "", "", "", "", "", ""],
        ["a. Penggantian Avtur untuk Mob - Demob", jam_mobdemob, "jam", jumlah_liter, "liter", harga_avtur_per_liter, biaya_penggantian_avtur_pesawat],
        ["b. Selama Operasi", jam_terbang_total, "jam", jumlah_liter, "liter", harga_avtur_per_liter, biaya_alutsista_pesawat_selama_operasi],
        ["c. Modifikasi dan inspeksi pesawat before - after rain making ", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_paket_modifikasi_inspeksi, biaya_modifikasi_dan_inspeksi_pesawat],
    ]

if jenis_pesawat == "Sewa Pesawat Swasta":
    data += [
        ["2. Sewa Pesawat", "", "", "", "", "", ""],
        ["a. Mobilisasi - Demobilisasi Pesawat", jam_mobdemob, "jam", p, "P", harga_sewa_per_jam, harga_mobdemob],
        ["b. Selama Operasi", jam_terbang_operasi, "jam", jh_keseluruhan, "hari", harga_sewa_per_jam, biaya_harga_sewa_pesawat],
    ]

data += [
    ["3. Kebutuhan Operasional Lapangan", "", "", "", "", "", ""],
    ["a. Sewa Kendaraan", "", "", "", "", "", ""],
    ["     Sebelum Operasi", unit_kendaraan_sebelum_operasi, "unit", jh_sewa_kendaraan_sebelum_setelah_operasi, "hari", sewa_kendaraan, biaya_sewa_kendaraan_sebelum_operasi],
    ["     Selama Operasi", jumlah_unit_kendaraan_keseluruhan, "unit", jh_keseluruhan, "hari", sewa_kendaraan, biaya_sewa_kendaraan_selama_operasi],
    ["     Setelah Operasi", unit_kendaraan_setelah_operasi, "unit", jh_sewa_kendaraan_sebelum_setelah_operasi, "hari", sewa_kendaraan, biaya_sewa_kendaraan_setelah_operasi],
    ["b. Peralatan dan Pendukung lapangan (ATK dan  Rompi Safety)", jumlah_paket, "paket", jh_keseluruhan, "hari", harga_peralatan_dan_pendukung_lapangan, biaya_peralatan_dan_pendukung_lapangan],
    ["c. Sewa Ruangan untuk Posko dan Gudang Bahan Semai", jumlah_paket, "paket", jh_keseluruhan, "hari", harga_paket_sewa_gedung, biaya_sewa_gedung],
    ["d. Ekspedisi", "", "", "", "", "", ""],
    ["     Peralatan", total_berat_peralatan, "kg", pp, "PP", biaya_kirim, biaya_ekspedisi_peralatan],
    ["     Bahan Semai", total_berat_bahan_semai, "kg", p, "P", biaya_kirim, biaya_ekspedisi_bahan_semai],
    ["e. Dukungan teknis operasional di lapangan", jumlah_paket, "orang", jh_keseluruhan, "hari", uang_harian_dukungan_teknis, biaya_dukungan_teknis_operasional_di_lapangan],
    ["f. Stanby  Air mobil  Pemadam Kebakaran dan untuk Cuci Pesawat", jumlah_paket, "paket", jh_keseluruhan, "hari", harga_air_mobil_pk_cuci_pesawat, biaya_air_mobil_pk_cuci_pesawat],
    ["C. Pelaporan", "", "", "", "", "", ""],
    ["g. Groundhandlng ( Forklift, Ground Support Equipment)", jumlah_paket, "paket", jh_keseluruhan, "hari", harga_groundhandling, biaya_groundhandling],
    ["Pencetakan dan Penggandaan Laporan", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_pencetakan_dan_penggandaan_laporan, biaya_pencetakan_dan_penggandaan_laporan],
]

# hitung total biaya (kolom paling kanan)
jumlah_total_biaya_omc = sum(row[-1] for row in data if row[-1] != "")

df_out = pd.DataFrame(data, columns=[
    "Kegiatan", "Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume", "Indeks (Rupiah)", "Biaya (Rupiah)"
])

# Simpan ke Excel

jumlah_harian = jumlah_total_biaya_omc / jh_keseluruhan

# buat baris ringkasan
summary_rows = pd.DataFrame([
    ["Jumlah", "", "", "", "", f"Total biaya {jh_keseluruhan} hari", jumlah_total_biaya_omc],
    ["", "", "", "", "", "Tarif Harian", jumlah_harian]
], columns=df_out.columns)

# gabungkan tabel utama dengan ringkasan
df_final = pd.concat([df_out, summary_rows], ignore_index=True)

# format angka pakai koma ribuan
# pastikan dulu dua kolom angka jadi numeric
df_final["Indeks (Rupiah)"] = pd.to_numeric(df_final["Indeks (Rupiah)"], errors="coerce")
df_final["Biaya (Rupiah)"]  = pd.to_numeric(df_final["Biaya (Rupiah)"], errors="coerce")

# lalu format dengan ribuan pakai koma
df_final["Indeks (Rupiah)"] = df_final["Indeks (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)
df_final["Biaya (Rupiah)"] = df_final["Biaya (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)

# st.write("📌 Rincian Biaya:", jumlah_total_biaya_omc)

# tampilkan dataframe di streamlit
# tampilkan tabel dengan styling
st.markdown(
    """
    <style>
    .dataframe td, .dataframe th {
        white-space: nowrap;
        text-align: left;
        padding: 8px 12px;
    }
    .dataframe th {
        background-color: #f5f5f5;
    }
    .dataframe td:nth-child(1) { min-width: 300px; }  /* kolom Uraian */
    .dataframe td:nth-child(2) { min-width: 50px; text-align: right; } /* kolom Jumlah */
    .dataframe td:nth-child(7) { min-width: 120px; text-align: right; } /* kolom Biaya */
    </style>
    """,
    unsafe_allow_html=True
)

st.dataframe(df_final, use_container_width=True)

from openpyxl.styles import Alignment, Font, numbers, Border, Side

# simpan ke buffer, bukan file fisik
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Laporan")

    ws = writer.sheets["Laporan"]

    # auto-width kolom
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # header bold & rata tengah + border
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border
        # semua header rata tengah
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- isi tabel ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value  # ambil nama header
            if col_name in ["Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

# reset pointer buffer
buffer.seek(0)

# tombol download
st.download_button(
    label="📥 Download Excel",
    data=buffer,
    file_name=f"Rincian Biaya Jasa Operasi Modifikasi Cuaca Provinsi {provinsi}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
st.markdown("---", unsafe_allow_html=True)
st.markdown(
    """
    <style>
        .footer-container {
            background-color: #f0f2f6;
            padding: 20px 40px;
            font-size: 14px;
            color: #333;
            display: flex;
            flex-wrap: wrap;
            justify-content: space-between;
        }
        .footer-column {
            flex: 1;
            min-width: 200px;
            margin-bottom: 20px;
        }
        .footer-title {
            font-weight: bold;
            margin-bottom: 10px;
            color: #0059b3;
        }
        .footer-bottom {
            text-align: center;
            font-size: 13px;
            padding: 10px;
            color: gray;
            border-top: 1px solid #ccc;
            background-color: #f0f2f6;
        }
    </style>

    <div class="footer-container">
        <div class="footer-column">
            <div class="footer-title">Kontak Kami</div>
            Jl. Angkasa I No.2 Kemayoran,<br>
            Jakarta Pusat 10610,<br>
            PO Box 3540 Jkt.<br><br>
            Contact Center: (021) 4246321<br>
            Faks: (021) 4246703<br>
            Email: tu5@bmkg.go.id
        </div>
        <div class="footer-column">
            <div class="footer-title">Media Sosial</div>
            Instagram : dmc_bmkg<br>
        </div>
    </div>

    <div class="footer-bottom">
        © 2025 - Sub Bagian Tata Usaha Modifikasi Cuaca, Deputi Bidang Modifikasi Cuaca, BMKG.
    </div>
    """,
    unsafe_allow_html=True
)
