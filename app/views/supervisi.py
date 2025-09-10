import streamlit as st
from utils import *
import datetime
import io
import pandas as pd

# === Judul App ===
st.markdown(
    """
    <h1 style='
        text-align: center;
        font-size: 45px;
        font-weight: bold;
        background: -webkit-linear-gradient(45deg, #1ABC9C, #3498DB);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    '>
        Sistem Layanan Modifikasi Cuaca
    </h1>
    """,
    unsafe_allow_html=True
)
# === Form Input Data Baru ===
st.subheader("Layanan Jasa Supervisi Operasi Modifikasi Cuaca")

# =================================================== #
# ==================== Baca Data ==================== #
# =================================================== #
df = pd.read_excel("database/data_gabungan_sdm2.xlsx", header=1) 

# ==================================================== #
# ==================== Input Data ==================== #
# ==================================================== #
today = datetime.date.today()

# hanya pilih tanggal mulai, bukan range
start_date = st.date_input(
    "Tanggal Mulai Pelaksanaan ",
    today,
    min_value=today,
    format="DD.MM.YYYY",
)
jumlah_input_hari=st.number_input("Masukkan Jumlah Hari :  ", step=1)
# otomatis tentukan end_date supaya total 7 hari
end_date = start_date + datetime.timedelta(days=jumlah_input_hari)

jh_keseluruhan=jumlah_input_hari+2

st.info(f"Pelaksanaan: {start_date} s/d {end_date} ({jh_keseluruhan} hari)")
# === Input Provinsi ===
provinsi = st.selectbox(
    "Tempat Pelaksanaan (Provinsi)",
    [
        "Aceh","Sumatera Utara","Sumatera Barat","Riau","Jambi","Sumatera Selatan",
        "Bengkulu","Lampung","Bangka Belitung","Kepulauan Riau","DKI Jakarta",
        "Jawa Barat","Jawa Tengah","DI Yogyakarta","Jawa Timur","Banten","Bali",
        "Nusa Tenggara Barat","Nusa Tenggara Timur","Kalimantan Barat","Kalimantan Tengah",
        "Kalimantan Selatan","Kalimantan Timur","Kalimantan Utara","Sulawesi Utara",
        "Sulawesi Tengah","Sulawesi Selatan","Sulawesi Tenggara","Gorontalo",
        "Sulawesi Barat","Maluku","Maluku Utara","Papua","Papua Barat","Papua Tengah",
        "Papua Pegunungan","Papua Selatan","Papua Barat Daya"
    ]
)

jenis_operasi = st.selectbox(

    "Pilih Lama Operasi",

    ["", "12 jam", "24 jam"]  # "" supaya default kosong

)

jumlah_personel_keseluruhan = 4


if jenis_operasi == "24 jam": 

    jumlah_personel_keseluruhan = 8


# Jumlah Paket Secara Umum
jumlah_paket = 1

# Jumlah Kali Paket
jumlah_kali_paket = 1

# Ambil uang harian berdasarkan provinsi
row = df[df['PROVINSI'] == provinsi]
if row.empty:
    st.error(f"Provinsi '{provinsi}' tidak ditemukan di data Excel!")
    st.stop()

# ===================================================== #
# ==================== Uang Harian ==================== #
# ===================================================== #
uang_harian_luar_kota = row['luar_kota'].values[0]

# Luar Kota

uh_tim_pelaksana = hitung_uang_harian(uang_harian_luar_kota, jh_keseluruhan, jumlah_personel_keseluruhan)


# ========================================================== #
# ==================== Biaya Penginapan ==================== #
# ========================================================== #


harga_penginapan_ktg3 = row['HOTELIII'].values[0]
harga_penginapan_ktg4 = row['HOTELIV'].values[0]
harga_penginapan_khusus = df[df['PROVINSI'] == 'Jawa Barat']['HOTELIV'].values[0]

# Biaya Penginapan Kategori III
# salah
biaya_penginapan_tim_pelaksana = biaya_penginapan(harga_penginapan_ktg4, jh_keseluruhan, jumlah_personel_keseluruhan)

# ===================================================== #
# ==================== Biaya Tiket ==================== #
# ===================================================== #

pp = 1
p = 1
biaya_tiket = row['PESAWAT_EKONOMI'].values[0] # Biaya tiket sama semua


biaya_tiket_tim_pelaksana = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_keseluruhan)

# ===================================================== #
# ==================== Harga Taksi ==================== #
# ===================================================== #

# jasa_omc.py
harga_taksi_flat = 2 * df[df['PROVINSI'] == 'DKI Jakarta']['BANDARA'].values[0]


biaya_taksi_tim_pelaksana = biaya_taksi(jumlah_personel_keseluruhan, harga_taksi_flat)

# =================== Honor Expertise ==================== #
honor_expertise = 350000

# Fungsi Biaya Honor Expertise
def biaya_honor_expertise(honor_expertise, jumlah_personel_keseluruhan, jh_keseluruhan):
    jumlah_biaya_honor_expertise = jumlah_personel_keseluruhan * honor_expertise * jh_keseluruhan
    # print(f"Jumlah biaya tiket untuk {provinsi}: Rp {jumlah_biaya_tiket_pelaksanaan:,}")
    return jumlah_biaya_honor_expertise

jumlah_biaya_honor_expertise = biaya_honor_expertise(honor_expertise, jumlah_personel_keseluruhan, jh_keseluruhan)
# ========================================================== #
# ==================== Sarana Prasarana ==================== #
# ========================================================== #

# =================== Data dan Informasi Cuaca ==================== #
data_informasi_cuaca = 4200000


# ========== Sewa Kendaraan ========== #
# Variabel sewa kendaraan
sewa_kendaraan = row['MOBIL'].values[0]
jumlah_unit_kendaraan_keseluruhan = math.ceil(jumlah_personel_keseluruhan / 4) 

# Biaya Sewa
biaya_sewa_kendaraan_selama_operasi = biaya_sewa_kendaraan(jh_keseluruhan, jumlah_unit_kendaraan_keseluruhan, sewa_kendaraan)


# ========== C. Pelaporan ========== #
# ========== Pencetakan dan Penggandaan Laporan ========== #
harga_pencetakan_dan_penggandaan_laporan = 10000000
biaya_pencetakan_dan_penggandaan_laporan = pencetakan_dan_penggandaan_laporan(jumlah_paket, jumlah_kali_paket, harga_pencetakan_dan_penggandaan_laporan)

# === Debug / Tampilkan hasil ===
# st.write("📌 Ringkasan Input:")
# st.write("Jumlah Hari:", jh_keseluruhan)
# st.write("Provinsi:", provinsi)
# st.write("Jenis Operasi:", jenis_operasi)
# st.write("Jumlah Personel OMC:", jumlah_personel_keseluruhan)

# Buat tabel hasil
data = [
    ["Provinsi " + provinsi, "", "", "", "", "", ""],
    ["A. Personil Pelaksana", "", "", "", "", "", ""],
    ["1. Uang Harian", jumlah_personel_keseluruhan, "orang", jh_keseluruhan, "hari", uang_harian_luar_kota, uh_tim_pelaksana],
    ["2. Biaya Penginapan", jumlah_personel_keseluruhan, "orang", jh_keseluruhan, "hari", harga_penginapan_ktg4, biaya_penginapan_tim_pelaksana],
    ["3. Biaya Tiket", jumlah_personel_keseluruhan, "orang", pp, "pp", biaya_tiket, biaya_tiket_tim_pelaksana],
    ["4. Taksi", jumlah_personel_keseluruhan, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_tim_pelaksana],
    ["5. Honor Expertise", jumlah_personel_keseluruhan, "orang", jh_keseluruhan, "hari", honor_expertise, jumlah_biaya_honor_expertise],

    ["B. Sarana dan Prasarana", "", "", "", "", "", ""],
    ["a. Data dan Informasi Cuaca ", jumlah_paket, "paket", jumlah_kali_paket, "kali", data_informasi_cuaca, data_informasi_cuaca],
    ["b. Sewa Kendaraan", jumlah_unit_kendaraan_keseluruhan, "unit", jumlah_input_hari, "hari", sewa_kendaraan,  biaya_sewa_kendaraan_selama_operasi],
   
    ["C. Hasil Akhir Supervisi Operasi Modifikasi Cuaca", "", "", "", "", "", ""],
    ["Laporan Supervisi Pelaksanaan Operasi Modifikasi Cuaca", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_pencetakan_dan_penggandaan_laporan, biaya_pencetakan_dan_penggandaan_laporan],
]

# hitung total biaya (kolom paling kanan)
jumlah_total_biaya_omc = sum(row[-1] for row in data if row[-1] != "")

df_out = pd.DataFrame(data, columns=[
    "Kegiatan", "Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume", "Indeks (Rupiah)", "Biaya (Rupiah)"
])

# Simpan ke Excel

jumlah_harian = jumlah_total_biaya_omc / jumlah_input_hari

# buat baris ringkasan
summary_rows = pd.DataFrame([
    ["Jumlah", "", "", "", "", f"Total biaya {jumlah_input_hari} hari", jumlah_total_biaya_omc],
    ["", "", "", "", "", "Tarif Harian", jumlah_harian]
], columns=df_out.columns)

# gabungkan tabel utama dengan ringkasan
df_final = pd.concat([df_out, summary_rows], ignore_index=True)

# format angka pakai koma ribuan
# pastikan dulu dua kolom angka jadi numeric
# df_final["Indeks (Rupiah)"] = pd.to_numeric(df_final["Indeks (Rupiah)"], errors="coerce")
df_final["Biaya (Rupiah)"]  = pd.to_numeric(df_final["Biaya (Rupiah)"], errors="coerce")

# lalu format dengan ribuan pakai koma
# df_final["Indeks (Rupiah)"] = df_final["Indeks (Rupiah)"].apply(
#     lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
# )
df_final["Biaya (Rupiah)"] = df_final["Biaya (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)

# st.write("📌 Rincian Biaya:", jumlah_total_biaya_omc)

# tampilkan dataframe di streamlit
# tampilkan tabel dengan styling
st.markdown(
    """
    <style>
    .dataframe td, .dataframe th {
        white-space: nowrap;
        text-align: left;
        padding: 8px 12px;
    }
    .dataframe th {
        background-color: #f5f5f5;
    }
    .dataframe td:nth-child(1) { min-width: 300px; }  /* kolom Uraian */
    .dataframe td:nth-child(2) { min-width: 50px; text-align: right; } /* kolom Jumlah */
    .dataframe td:nth-child(7) { min-width: 120px; text-align: right; } /* kolom Biaya */
    </style>
    """,
    unsafe_allow_html=True
)

st.dataframe(df_final, use_container_width=True)

from openpyxl.styles import Alignment, Font, numbers, Border, Side

# simpan ke buffer, bukan file fisik
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Laporan")

    ws = writer.sheets["Laporan"]

    # auto-width kolom
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # header bold & rata tengah + border
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border
        # semua header rata tengah
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- isi tabel ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value  # ambil nama header
            if col_name in ["Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

# reset pointer buffer
buffer.seek(0)

# tombol download
st.download_button(
    label="📥 Download Excel",
    data=buffer,
    file_name=f"Rincian Biaya Jasa Operasi Modifikasi Cuaca Provinsi {provinsi}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
st.markdown("---", unsafe_allow_html=True)
st.markdown(
    """
    <style>
        .footer-container {
            background-color: #f0f2f6;
            padding: 20px 40px;
            font-size: 14px;
            color: #333;
            display: flex;
            flex-wrap: wrap;
            justify-content: space-between;
        }
        .footer-column {
            flex: 1;
            min-width: 200px;
            margin-bottom: 20px;
        }
        .footer-title {
            font-weight: bold;
            margin-bottom: 10px;
            color: #0059b3;
        }
        .footer-bottom {
            text-align: center;
            font-size: 13px;
            padding: 10px;
            color: gray;
            border-top: 1px solid #ccc;
            background-color: #f0f2f6;
        }
    </style>

    <div class="footer-container">
        <div class="footer-column">
            <div class="footer-title">Kontak Kami</div>
            Jl. Angkasa I No.2 Kemayoran,<br>
            Jakarta Pusat 10610,<br>
            PO Box 3540 Jkt.<br><br>
            Contact Center: (021) 4246321<br>
            Faks: (021) 4246703<br>
            Email: tu5@bmkg.go.id
        </div>
        <div class="footer-column">
            <div class="footer-title">Media Sosial</div>
            Instagram : dmc_bmkg<br>
        </div>
    </div>

    <div class="footer-bottom">
        © 2025 - Sub Bagian Tata Usaha Modifikasi Cuaca, Deputi Bidang Modifikasi Cuaca, BMKG.
    </div>
    """,
    unsafe_allow_html=True
)
