import streamlit as st
from utils import *
import datetime
import io
import pandas as pd

# === Form Input Data Baru ===
st.subheader("Biaya Layanan Operasional Modifikasi Cuaca Berbasis Wahana Pesawat Nirawak")

# =================================================== #
# ==================== Baca Data ==================== #
# =================================================== #
df = pd.read_excel("database/data_gabungan_sdm1.xlsx", header=1) 

# ==================================================== #
# ==================== Input Data ==================== #
# ==================================================== #
today = datetime.date.today()

# hanya pilih tanggal mulai, bukan range
tanggal_range = st.date_input(
    "Tanggal Pelaksanaan",
    (today, today + datetime.timedelta(days=6)),  # default 7 hari
    min_value=today,
    format="DD.MM.YYYY",
)

# Hitung jumlah hari
jh_keseluruhan = None
if isinstance(tanggal_range, tuple) and len(tanggal_range) == 2:
    start_date, end_date = tanggal_range
    jh_keseluruhan = (end_date - start_date).days + 1
    st.info(f"Pelaksanaan: {start_date} s/d {end_date} ({jh_keseluruhan} hari)")

# === Input Provinsi ===
provinsi = st.selectbox(
    "Tempat Pelaksanaan (Provinsi)",
    [
        "Aceh","Sumatera Utara","Sumatera Barat","Riau","Jambi","Sumatera Selatan",
        "Bengkulu","Lampung","Kepulauan Bangka Belitung","Kepulauan Riau","DKI Jakarta",
        "Jawa Barat","Jawa Tengah","DI Yogyakarta","Jawa Timur","Banten","Bali",
        "Nusa Tenggara Barat","Nusa Tenggara Timur","Kalimantan Barat","Kalimantan Tengah",
        "Kalimantan Selatan","Kalimantan Timur","Kalimantan Utara","Sulawesi Utara",
        "Sulawesi Tengah","Sulawesi Selatan","Sulawesi Tenggara","Gorontalo",
        "Sulawesi Barat","Maluku","Maluku Utara","Papua","Papua Barat","Papua Tengah",
        "Papua Pegunungan","Papua Selatan","Papua Barat Daya"
    ]
)

# === Input Jumlah Sistem Teleburning ===
jumlah_unit_pesawat_nirawak = st.number_input(
    "Jumlah Pesawat Nirawak",
    min_value=1, step=1, value=1
)

# ===================================================== #
# ==================== Variabel Fix =================== #
# ===================================================== #
# Jumlah Personel Lengkap
jumlah_personel_sebelum_operasi = 2 # FIX, edit jika berubah
jumlah_personel_selama_operasi = 13
jumlah_personel_sesudah_operasi = 2
jumlah_personel_lokal_polri = 3

# Ambil uang harian berdasarkan provinsi
row = df[df['PROVINSI'] == provinsi]
if row.empty:
    st.error(f"Provinsi '{provinsi}' tidak ditemukan di data Excel!")
    st.stop()

pp = 1

jumlah_paket = 1
jumlah_kali_paket = 1

jumlah_hari_paket = 1

# ===================================================== #
# ==================== Uang Harian ==================== #
# ===================================================== #
uang_harian_luar_kota = row['luar_kota'].values[0]
uang_harian_dalam_kota = row['dalam_kota'].values[0]
uang_harian_khusus = df[df['PROVINSI'] == 'Jawa Barat']['luar_kota'].values[0]

jh_uh_selama_operasi = jh_keseluruhan
jh_uh_sebelum_operasi = 2
jh_uh_sesudah_operasi = 2

# Luar Kota
uh_sebelum_operasi = hitung_uang_harian(uang_harian_luar_kota, jh_uh_sebelum_operasi, jumlah_personel_sebelum_operasi)
uh_selama_operasi = hitung_uang_harian(uang_harian_luar_kota, jh_uh_selama_operasi, jumlah_personel_selama_operasi)
uh_sesudah_operasi = hitung_uang_harian(uang_harian_luar_kota, jh_uh_sesudah_operasi, jumlah_personel_sesudah_operasi)

# ========================================================== #
# ==================== Biaya Penginapan ==================== #
# ========================================================== #
harga_penginapan_ktg3 = row['HOTELIII'].values[0]
harga_penginapan_ktg4 = row['HOTELIV'].values[0]
harga_penginapan_khusus = df[df['PROVINSI'] == 'Jawa Barat']['HOTELIV'].values[0]

jh_penginapan_sebelum_operasi = jh_uh_sebelum_operasi-1
jh_penginapan_selama_operasi = jh_uh_selama_operasi-1
jh_penginapan_sesudah_operasi = jh_uh_sesudah_operasi-1

# Biaya Penginapan Kategori III
biaya_penginapan_sebelum_operasi = biaya_penginapan(harga_penginapan_ktg3, jh_penginapan_sebelum_operasi, jumlah_personel_sebelum_operasi)
biaya_penginapan_selama_operasi = biaya_penginapan(harga_penginapan_ktg3, jh_penginapan_selama_operasi, jumlah_personel_selama_operasi)
biaya_penginapan_sesudah_operasi = biaya_penginapan(harga_penginapan_ktg3, jh_penginapan_sesudah_operasi, jumlah_personel_sesudah_operasi)

# ===================================================== #
# ==================== Biaya Tiket ==================== #
# ===================================================== #
biaya_tiket = row['PESAWAT_EKONOMI'].values[0] # Biaya tiket sama semua

biaya_tiket_sebelum_operasi = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_sebelum_operasi)
biaya_tiket_selama_operasi = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_selama_operasi)
biaya_tiket_sesudah_operasi = biaya_tiket_pelaksanaan(biaya_tiket, pp, jumlah_personel_sesudah_operasi)

# ===================================================== #
# ==================== Harga Taksi ==================== #
# ===================================================== #
harga_taksi_flat = 2 * df[df['PROVINSI'] == 'DKI Jakarta']['BANDARA'].values[0]

biaya_taksi_sebelum_operasi = biaya_taksi(jumlah_personel_sebelum_operasi, harga_taksi_flat)
biaya_taksi_selama_operasi = biaya_taksi(jumlah_personel_selama_operasi, harga_taksi_flat)
biaya_taksi_sesudah_operasi = biaya_taksi(jumlah_personel_sesudah_operasi, harga_taksi_flat)

# =================================================================== #
# ==================== Uang Harian Lokal (Polri) ==================== #
# =================================================================== #
uh_lokal_polri = 150000

biaya_uh_lokal_polri = tenaga_lokal(jumlah_personel_lokal_polri, jh_keseluruhan, uh_lokal_polri)

# ============================================================== #
# ==================== 1. Bahan Semai Flare ==================== #
# ============================================================== #
harga_bahan_semai_berbasis_nirawak = 2500000

jumlah_pcs, biaya_bahan_semai_berbasis_nirawak = bahan_semai_berbasis_nirawak(jumlah_unit_pesawat_nirawak, jh_keseluruhan, harga_bahan_semai_berbasis_nirawak)

# ================================================================= #
# ==================== 2. Sewa Pesawat Nirawak ==================== #
# ================================================================= #
harga_sewa_pesawat_nirawak = 5000000 # pindah ke jasa_omc_darat.py

biaya_sewa_pesawat_nirawak = sewa_pesawat_nirawak(jumlah_paket, jumlah_kali_paket, harga_sewa_pesawat_nirawak)

# ======================================================================== #
# ==================== 3. Perijinan Bahan Semai Flare ==================== #
# ======================================================================== #
harga_perijinan_bahan_semai_flare = 200000000

biaya_perijinan_bahan_semai_flare = perijinan_bahan_semai_flare(jumlah_paket, jumlah_kali_paket, harga_perijinan_bahan_semai_flare)

# ============================================================================ #
# ==================== biaya_sewa_kendaraan_survey_lokasi ==================== #
# ============================================================================ #
sewa_kendaraan = row['MOBIL'].values[0]
jumlah_unit_kendaraan_keseluruhan = math.ceil(jumlah_personel_selama_operasi / 4)

biaya_sewa_kendaraan_darat= biaya_sewa_kendaraan(jh_keseluruhan, jumlah_unit_kendaraan_keseluruhan, sewa_kendaraan) # pindah ke jasa_survey.py

# ========================================================================== #
# ==================== Peralatan dan Pendukung Lapangan ==================== #
# ========================================================================== #
harga_peralatan_dan_pendukung_lapangan = 4275000

biaya_peralatan_dan_pendukung_lapangan = peralatan_dan_pendukung_lapangan(jumlah_paket, jumlah_hari_paket, harga_peralatan_dan_pendukung_lapangan)

# ============================================================================================== #
# ==================== Laporan Hasil Survey Lokasi dan Instalasi Menara GBG ==================== #
# ============================================================================================== #
harga_pencetakan_dan_penggandaan_laporan = 10000000
biaya_pencetakan_dan_penggandaan_laporan = pencetakan_dan_penggandaan_laporan(jumlah_paket, jumlah_kali_paket, harga_pencetakan_dan_penggandaan_laporan)

# # === Debug / Tampilkan hasil ===
# st.write("📌 Ringkasan Input:")
# st.write("Jumlah Hari:", jh_keseluruhan)
# st.write("Provinsi:", provinsi)

# Buat tabel hasil
data = [
    ["Provinsi " + provinsi, "", "", "", "", "", ""],
    ["A. Personil Pelaksana", "", "", "", "", "", ""],

    ["1. Sebelum Operasi", "", "", "", "", "", ""],
    ["     Uang Harian", jumlah_personel_sebelum_operasi, "orang", jh_uh_sebelum_operasi, "hari", uang_harian_luar_kota, uh_sebelum_operasi],
    ["     Biaya Penginapan", jumlah_personel_sebelum_operasi, "orang", jh_penginapan_sebelum_operasi, "hari", harga_penginapan_ktg3, biaya_penginapan_sebelum_operasi],
    ["     Biaya Tiket ", jumlah_personel_sebelum_operasi, "orang", pp, "pp", biaya_tiket, biaya_tiket_sebelum_operasi],
    ["     Taksi", jumlah_personel_sebelum_operasi, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_sebelum_operasi],

    ["2. Selama Operasi", "", "", "", "", "", ""],
    ["     Uang Harian", jumlah_personel_selama_operasi, "orang", jh_uh_selama_operasi, "hari", uang_harian_luar_kota, uh_selama_operasi],
    ["     Biaya Penginapan", jumlah_personel_selama_operasi, "orang", jh_penginapan_selama_operasi, "hari", harga_penginapan_ktg4, biaya_penginapan_selama_operasi],
    ["     Biaya Tiket ", jumlah_personel_selama_operasi, "orang", pp, "pp", biaya_tiket, biaya_tiket_selama_operasi],
    ["     Taksi", jumlah_personel_selama_operasi, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_selama_operasi],
    ["     Honor Tenaga Lokal", jumlah_personel_lokal_polri, "kegiatan", "", "", uh_lokal_polri, biaya_uh_lokal_polri],

    ["3. Sesudah Operasi", "", "", "", "", "", ""],
    ["     Uang Harian", jumlah_personel_sesudah_operasi, "orang", jh_uh_sesudah_operasi, "hari", uang_harian_luar_kota, uh_sesudah_operasi],
    ["     Biaya Penginapan", jumlah_personel_sesudah_operasi, "orang", jh_penginapan_sesudah_operasi, "hari", harga_penginapan_ktg3, biaya_penginapan_sesudah_operasi],
    ["     Biaya Tiket ", jumlah_personel_sesudah_operasi, "orang", pp, "pp", biaya_tiket, biaya_tiket_sesudah_operasi],
    ["     Taksi", jumlah_personel_sesudah_operasi, "orang", pp, "pp", harga_taksi_flat, biaya_taksi_sesudah_operasi],

    ["B. Sarana dan Prasarana", "", "", "", "", "", ""],
    ["1. Bahan Semai Flare", jumlah_pcs, "pcs", jumlah_kali_paket, "hari", harga_bahan_semai_berbasis_nirawak, biaya_bahan_semai_berbasis_nirawak],
    ["2. Sewa Pesawat Nirawak", jumlah_pcs, "unit", jh_keseluruhan, "hari", harga_sewa_pesawat_nirawak, biaya_sewa_pesawat_nirawak],
    ["3. Perijinan Bahan Semai Flare", jumlah_paket, "paket", jumlah_hari_paket, "hari", harga_perijinan_bahan_semai_flare, biaya_perijinan_bahan_semai_flare],
    ["4. Kebutuhan Operasional Lapangan", "", "", "", "", "", ""],
    ["    a. Sewa  kendaraan ", jumlah_unit_kendaraan_keseluruhan, "unit", jh_keseluruhan+3, "hari", sewa_kendaraan, biaya_sewa_kendaraan_darat],
    ["    b. Peralatan dan pendukung lapangan", jumlah_paket, "paket", 1, "hari", sewa_kendaraan, biaya_peralatan_dan_pendukung_lapangan],
    
    ["C. Hasil Layanan Operasional Modifikasi Cuaca Berbasis Wahana Penyemaian Awan dari Darat", "", "", "", "", "", ""],
    ["Pencetakan dan Penggandaan Laporan", jumlah_paket, "paket", jumlah_kali_paket, "kali", harga_pencetakan_dan_penggandaan_laporan, biaya_pencetakan_dan_penggandaan_laporan],
]

jumlah_total_jasa_omc_nirawak = sum(row[-1] for row in data if row[-1] != "")

data.append(["Jumlah", "", "", "", "", "", jumlah_total_jasa_omc_nirawak])

df_out = pd.DataFrame(data, columns=[
    "Uraian", "Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume", "Indeks (Rupiah)", "Biaya (Rupiah)"
])

jumlah_harian = jumlah_total_jasa_omc_nirawak / jh_keseluruhan

# buat baris ringkasan
summary_rows = pd.DataFrame([
    ["Jumlah", "", "", "", "", f"Total biaya {jh_keseluruhan} hari", jumlah_total_jasa_omc_nirawak],
    ["", "", "", "", "", "Tarif Harian", jumlah_harian]
], columns=df_out.columns)

# gabungkan tabel utama dengan ringkasan
df_final = pd.concat([df_out, summary_rows], ignore_index=True)

# format angka pakai koma ribuan
# pastikan dulu dua kolom angka jadi numeric
df_final["Indeks (Rupiah)"] = pd.to_numeric(df_final["Indeks (Rupiah)"], errors="coerce")
df_final["Biaya (Rupiah)"]  = pd.to_numeric(df_final["Biaya (Rupiah)"], errors="coerce")

# lalu format dengan ribuan pakai koma
df_final["Indeks (Rupiah)"] = df_final["Indeks (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)
df_final["Biaya (Rupiah)"] = df_final["Biaya (Rupiah)"].apply(
    lambda x: "{:,.0f}".format(x) if pd.notnull(x) else ""
)

st.write("📌 Rincian Biaya:", jumlah_total_jasa_omc_nirawak)

# tampilkan dataframe di streamlit
# tampilkan tabel dengan styling
st.markdown(
    """
    <style>
    .dataframe td, .dataframe th {
        white-space: nowrap;
        text-align: left;
        padding: 8px 12px;
    }
    .dataframe th {
        background-color: #f5f5f5;
    }
    .dataframe td:nth-child(1) { min-width: 300px; }  /* kolom Uraian */
    .dataframe td:nth-child(2) { min-width: 50px; text-align: right; } /* kolom Jumlah */
    .dataframe td:nth-child(7) { min-width: 120px; text-align: right; } /* kolom Biaya */
    </style>
    """,
    unsafe_allow_html=True
)

st.dataframe(df_final, use_container_width=True)

from openpyxl.styles import Alignment, Font, numbers, Border, Side

# simpan ke buffer, bukan file fisik
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Laporan")

    ws = writer.sheets["Laporan"]

    # auto-width kolom
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # header bold & rata tengah + border
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border
        # semua header rata tengah
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- isi tabel ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value  # ambil nama header
            if col_name in ["Jumlah", "Satuan Jumlah", "Volume", "Satuan Volume"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

# reset pointer buffer
buffer.seek(0)

# tombol download
st.download_button(
    label="📥 Download Excel",
    data=buffer,
    file_name=f"Laporan Biaya Layanan Operasional Modifikasi Cuaca Berbasis Wahana Pesawat Nirawak Provinsi_{provinsi}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)